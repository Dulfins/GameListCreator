from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO


def _safe_float(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(str(x).strip())
    except:
        return None


def build_selected_games_workbook(games: list[dict]) -> BytesIO:
    """
    Build and return an Excel workbook (in-memory) for selected games.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Games Backlog"

    headers = ["Game", "Owned", "Excitement", "Hours to Complete", "Completed? (Y/N)", "Notes"]
    start_row = 3
    header_row = start_row
    first_data_row = header_row + 1
    n = len(games)
    last_data_row = first_data_row + n - 1 if n > 0 else first_data_row

    # ---- Summary ----
    ws["A1"] = "Total Hours"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = f"=SUM(D{first_data_row}:D{max(last_data_row, first_data_row)})"
    ws["B1"].number_format = "0.0"
    ws["B1"].font = Font(bold=True)

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ---- Data ----
    for g in games:
        ws.append([
            g.get("game_name", ""),
            "Y" if g.get("owned") else "N",
            g.get("intrigue", ""),
            _safe_float(g.get("main_extra")),
            "N",
            ""
        ])

    # ---- Formatting ----
    widths = [34, 10, 14, 18, 18, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{first_data_row}"

    for r in range(first_data_row, last_data_row + 1):
        ws[f"D{r}"].number_format = "0.0"
        ws[f"C{r}"].number_format = "0"
        ws[f"F{r}"].alignment = Alignment(wrap_text=True)

    # ---- Validation ----
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E{first_data_row}:E{max(last_data_row, first_data_row)}")

    # ---- Table ----
    if n > 0:
        table = Table(
            displayName="SelectedGames",
            ref=f"A{header_row}:F{last_data_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )
        ws.add_table(table)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
